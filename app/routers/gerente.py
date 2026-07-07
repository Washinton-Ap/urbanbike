"""Rutas para el rol Gerente — analítica sobre ClickHouse y gestión de usuarios."""

import io
import json
import urllib.parse
import urllib.request
from datetime import datetime, timezone

from fastapi import APIRouter, File, Form, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse

from app.db import clickhouse as ch
from app.db.pocketbase import get_admin_client, registrar_auditoria
from app.templating import templates

router = APIRouter(prefix="/gerente", tags=["gerente"])

DIAS = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
POR_PAGINA = 25


def _ctx(request: Request, **extra) -> dict:
    return {"user": getattr(request.state, "user", None), **extra}


def _pb():
    import app.db.pocketbase as pbmod
    try:
        return get_admin_client()
    except Exception:
        pbmod._admin_client = None
        return get_admin_client()


def _flash(request: Request, url: str, tipo: str, msg: str) -> RedirectResponse:
    request.session["flash"] = {"type": tipo, "msg": msg}
    return RedirectResponse(url, status_code=302)


_ACCION_TIPO = {"crear": "crear", "editar": "editar", "eliminar": "eliminar"}
_MODULO_PLURAL = {"bicicleta": "bicicletas", "estación": "estaciones", "tarifa": "tarifas", "usuario": "usuarios"}


def _log(request: Request, accion: str, detalle: str) -> None:
    """Registra una acción de CRUD del gerente en la bitácora de cambios y en la auditoría."""
    user = getattr(request.state, "user", {}) or {}
    try:
        _pb().create_record("bitacora_cambios", {
            "usuario_nombre": user.get("name") or user.get("email", "Gerente"),
            "accion":  accion,
            "detalle": detalle,
            "fecha":   datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        })
    except Exception:
        pass

    palabras = accion.lower().split(" ", 1)
    accion_tipo = _ACCION_TIPO.get(palabras[0], "editar")
    modulo = _MODULO_PLURAL.get(palabras[1], "sistema") if len(palabras) > 1 else "sistema"
    registrar_auditoria(
        user.get("pb_token", ""), user.get("id", ""),
        user.get("name") or user.get("email", "Gerente"), user.get("email", ""),
        accion_tipo, modulo, detalle, request,
        usuario_rol=user.get("rol_nombre") or user.get("rol_slug", ""),
    )


def _build_where(fecha_inicio: str, fecha_fin: str, membresia: str, tipo_bici: str) -> str:
    parts = [
        f"fecha_inicio >= toDateTime('{fecha_inicio} 00:00:00')",
        f"fecha_inicio <= toDateTime('{fecha_fin} 23:59:59')",
    ]
    if membresia == "member":
        parts.append("id_membresia = 2")
    elif membresia == "casual":
        parts.append("id_membresia = 1")
    if tipo_bici == "classic_bike":
        parts.append("id_tipo_bicicleta = 1")
    elif tipo_bici == "electric_bike":
        parts.append("id_tipo_bicicleta = 2")
    return "WHERE " + " AND ".join(parts)


@router.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    flash = request.session.pop("flash", None)
    kpis: dict = {}
    ch_ok = True
    dow_labels: list = []
    dow_values: list = []
    top_labels: list = []
    top_values: list = []
    try:
        kpis = ch.query_one("""
            SELECT
              count()                            AS total_viajes,
              countDistinct(id_estacion_inicio)   AS total_estaciones,
              round(avg(duracion_min), 2)         AS dur_prom_min,
              round(avg(distancia_km), 2)         AS dist_prom_km,
              countIf(id_membresia = 2)           AS viajes_member,
              countIf(id_membresia = 1)           AS viajes_casual,
              countIf(id_tipo_bicicleta = 1)      AS viajes_clasica,
              countIf(id_tipo_bicicleta = 2)      AS viajes_electrica
            FROM fact_viajes
        """) or {}

        dow_rows = {r["dia"]: r["viajes"] for r in ch.query("""
            SELECT toDayOfWeek(fecha_inicio) AS dia, count() AS viajes
            FROM fact_viajes GROUP BY dia ORDER BY dia
        """)}
        dow_labels = DIAS
        dow_values = [dow_rows.get(i, 0) for i in range(1, 8)]

        top10 = ch.query("""
            SELECT e.nombre_estacion AS nombre, count() AS viajes
            FROM fact_viajes f
            LEFT JOIN dim_estaciones e ON f.id_estacion_inicio = e.id_estacion
            GROUP BY e.nombre_estacion ORDER BY viajes DESC LIMIT 10
        """)
        top_labels = [str(r.get("nombre") or "N/A")[:30] for r in top10]
        top_values = [r["viajes"] for r in top10]

    except Exception:
        ch_ok = False

    return templates.TemplateResponse(request, "gerente/dashboard.html", _ctx(request,
        title="Dashboard — Gerente", flash=flash, kpis=kpis, ch_ok=ch_ok,
        dow_labels=json.dumps(dow_labels),
        dow_values=json.dumps(dow_values),
        top_labels=json.dumps(top_labels),
        top_values=json.dumps(top_values),
    ))


@router.get("/reportes", response_class=HTMLResponse)
def reportes(
    request: Request,
    fecha_inicio: str = Query("2023-10-01"),
    fecha_fin: str = Query("2023-10-31"),
    membresia: str = Query("all"),
    tipo_bici: str = Query("all"),
    pagina: int = Query(1, ge=1),
):
    flash = request.session.pop("flash", None)
    where = _build_where(fecha_inicio, fecha_fin, membresia, tipo_bici)
    offset = (pagina - 1) * POR_PAGINA

    estaciones: list[dict] = []
    total_viajes = 0
    total_filas = 0
    total_paginas = 1
    chart_est_labels: list = []
    chart_est_values: list = []
    chart_dow_values: list = []
    chart_tipo_labels: list = []
    chart_tipo_values: list = []
    chart_tendencia_labels: list = []
    chart_tendencia_values: list = []
    ch_ok = True

    try:
        cnt = ch.query_one(f"""
            SELECT countDistinct(id_estacion_inicio) AS total_est,
                   count() AS total_viajes
            FROM fact_viajes {where}
        """)
        total_filas = cnt.get("total_est", 0) if cnt else 0
        total_viajes = cnt.get("total_viajes", 0) if cnt else 0
        total_paginas = max(1, (total_filas + POR_PAGINA - 1) // POR_PAGINA)

        estaciones = ch.query(f"""
            SELECT
              e.nombre_estacion           AS nombre,
              count()                     AS viajes,
              round(avg(f.duracion_min), 1) AS dur_prom,
              round(avg(f.distancia_km), 2) AS dist_prom
            FROM fact_viajes f
            LEFT JOIN dim_estaciones e ON f.id_estacion_inicio = e.id_estacion
            {where}
            GROUP BY e.nombre_estacion
            ORDER BY viajes DESC
            LIMIT {POR_PAGINA} OFFSET {offset}
        """)

        top10 = ch.query(f"""
            SELECT e.nombre_estacion AS nombre, count() AS viajes
            FROM fact_viajes f
            LEFT JOIN dim_estaciones e ON f.id_estacion_inicio = e.id_estacion
            {where}
            GROUP BY e.nombre_estacion ORDER BY viajes DESC LIMIT 10
        """)
        chart_est_labels = [str(r.get("nombre") or "N/A")[:28] for r in top10]
        chart_est_values = [r["viajes"] for r in top10]

        dow_rows = {r["dia"]: r["viajes"] for r in ch.query(f"""
            SELECT toDayOfWeek(fecha_inicio) AS dia, count() AS viajes
            FROM fact_viajes {where} GROUP BY dia ORDER BY dia
        """)}
        chart_dow_values = [dow_rows.get(i, 0) for i in range(1, 8)]

        tipo_rows = ch.query(f"""
            SELECT t.nombre AS nombre, round(avg(f.duracion_min), 1) AS dur_prom
            FROM fact_viajes f
            LEFT JOIN dim_tipos_bicicleta t ON f.id_tipo_bicicleta = t.id_tipo
            {where}
            GROUP BY t.nombre ORDER BY t.nombre
        """)
        chart_tipo_labels = [str(r.get("nombre") or "N/A") for r in tipo_rows]
        chart_tipo_values = [r["dur_prom"] for r in tipo_rows]

        tendencia_rows = ch.query(f"""
            SELECT toDate(fecha_inicio) AS dia, count() AS viajes
            FROM fact_viajes {where}
            GROUP BY dia ORDER BY dia
        """)
        chart_tendencia_labels = [r["dia"].strftime("%d/%m") if hasattr(r["dia"], "strftime") else str(r["dia"]) for r in tendencia_rows]
        chart_tendencia_values = [r["viajes"] for r in tendencia_rows]

    except Exception:
        ch_ok = False

    return templates.TemplateResponse(request, "gerente/reportes.html", _ctx(request,
        title="Reportes — Gerente", flash=flash, ch_ok=ch_ok,
        estaciones=estaciones,
        total_viajes=total_viajes, total_filas=total_filas,
        total_paginas=total_paginas, pagina=pagina,
        fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
        membresia=membresia, tipo_bici=tipo_bici,
        chart_est_labels=json.dumps(chart_est_labels),
        chart_est_values=json.dumps(chart_est_values),
        chart_dow_labels=json.dumps(DIAS),
        chart_dow_values=json.dumps(chart_dow_values),
        chart_tipo_labels=json.dumps(chart_tipo_labels),
        chart_tipo_values=json.dumps(chart_tipo_values),
        chart_tendencia_labels=json.dumps(chart_tendencia_labels),
        chart_tendencia_values=json.dumps(chart_tendencia_values),
    ))


@router.get("/reportes/excel")
def export_excel(
    request: Request,
    fecha_inicio: str = Query("2023-10-01"),
    fecha_fin: str = Query("2023-10-31"),
    membresia: str = Query("all"),
    tipo_bici: str = Query("all"),
):
    from datetime import datetime

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    where = _build_where(fecha_inicio, fecha_fin, membresia, tipo_bici)
    rows = ch.query(f"""
        SELECT
          e.nombre_estacion             AS nombre,
          count()                       AS viajes,
          round(avg(f.duracion_min), 1) AS dur_prom,
          round(avg(f.distancia_km), 2) AS dist_prom
        FROM fact_viajes f
        LEFT JOIN dim_estaciones e ON f.id_estacion_inicio = e.id_estacion
        {where}
        GROUP BY e.nombre_estacion ORDER BY viajes DESC LIMIT 1000
    """)

    BLUE = "1E86BD"
    WHITE = "FFFFFF"
    ALT   = "D6EDF8"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Estaciones"

    # Título principal
    ws.merge_cells("A1:D1")
    ws["A1"] = "UrbanBike — Reporte Analítico"
    ws["A1"].font = Font(name="Calibri", bold=True, color=BLUE, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtítulo con filtros
    mem_map  = {"all": "Todos", "member": "Miembros", "casual": "Casuales"}
    bici_map = {"all": "Todos", "classic_bike": "Clásica", "electric_bike": "Eléctrica"}
    ws.merge_cells("A2:D2")
    ws["A2"] = (
        f"Período: {fecha_inicio} → {fecha_fin}  |  "
        f"Membresía: {mem_map.get(membresia, 'Todos')}  |  "
        f"Tipo: {bici_map.get(tipo_bici, 'Todos')}  |  "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws["A2"].font = Font(name="Calibri", color="64748B", size=9)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    # Cabeceras de columna
    hdr_fill = PatternFill("solid", fgColor=BLUE)
    hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    center   = Alignment(horizontal="center", vertical="center")
    right    = Alignment(horizontal="right",  vertical="center")
    thin_bot = Border(bottom=Side(style="thin", color="E2E8F0"))

    for col, label in enumerate(
        ["Estación", "Viajes", "Duración Prom. (min)", "Distancia Prom. (km)"], start=1
    ):
        c = ws.cell(row=4, column=col, value=label)
        c.font = c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
    ws.row_dimensions[4].height = 20

    alt_fill = PatternFill("solid", fgColor=ALT)
    dat_font = Font(name="Calibri", size=10)

    for ri, row in enumerate(rows, start=5):
        nombre_c = ws.cell(row=ri, column=1, value=row.get("nombre") or "N/A")
        nombre_c.font = dat_font

        viajes_c = ws.cell(row=ri, column=2, value=int(row.get("viajes", 0)))
        viajes_c.font = dat_font
        viajes_c.number_format = "#,##0"
        viajes_c.alignment = right

        dur_c = ws.cell(row=ri, column=3, value=float(row.get("dur_prom", 0)))
        dur_c.font = dat_font
        dur_c.number_format = "0.0"
        dur_c.alignment = right

        dist_c = ws.cell(row=ri, column=4, value=float(row.get("dist_prom", 0)))
        dist_c.font = dat_font
        dist_c.number_format = "0.00"
        dist_c.alignment = right

        if ri % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=ri, column=col).fill = alt_fill
        for col in range(1, 5):
            ws.cell(row=ri, column=col).border = thin_bot

    # Fila de totales
    total_row = len(rows) + 5
    tot_font = Font(name="Calibri", bold=True, size=10, color=BLUE)
    ws.cell(row=total_row, column=1, value=f"Total: {len(rows)} estaciones").font = tot_font
    total_v = ws.cell(row=total_row, column=2, value=sum(int(r.get("viajes", 0)) for r in rows))
    total_v.font = tot_font
    total_v.number_format = "#,##0"
    total_v.alignment = right

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"urbanbike_reporte_{fecha_inicio}_{fecha_fin}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Reportes de Pagos ─────────────────────────────────────────────────────────

ESTADOS_PAGO = ["pendiente", "pagado", "verificacion_pendiente", "pendiente_efectivo", "rechazado", "cancelado"]
METODOS_PAGO = ["efectivo", "tarjeta", "transferencia"]


def _pagos_pb() -> list[dict]:
    return get_admin_client().list_records("pagos", sort="-fecha_pago", per_page=2000).get("items", [])


def _filtrar_pagos(pagos: list[dict], estado: str, metodo: str, fecha_inicio: str, fecha_fin: str) -> list[dict]:
    out = []
    for p in pagos:
        if estado != "all" and p.get("estado") != estado:
            continue
        if metodo != "all" and p.get("metodo_pago") != metodo:
            continue
        fecha = (p.get("fecha_pago") or "")[:10]
        if fecha_inicio and fecha and fecha < fecha_inicio:
            continue
        if fecha_fin and fecha and fecha > fecha_fin:
            continue
        out.append(p)
    return out


@router.get("/reportes/pagos", response_class=HTMLResponse)
def reportes_pagos(
    request: Request,
    estado: str = Query("all"),
    metodo: str = Query("all"),
    fecha_inicio: str = Query(""),
    fecha_fin: str = Query(""),
):
    from datetime import datetime, timezone
    import calendar

    flash = request.session.pop("flash", None)
    pb_ok = True
    pagos: list[dict] = []
    try:
        pagos = _pagos_pb()
    except Exception:
        pb_ok = False

    ahora = datetime.now(timezone.utc)
    hoy_str = ahora.strftime("%Y-%m-%d")
    mes_str = ahora.strftime("%Y-%m")
    dias_mes = calendar.monthrange(ahora.year, ahora.month)[1]

    pagados = [p for p in pagos if p.get("estado") == "pagado"]

    ingresos_dia = sum(float(p.get("monto_total") or 0) for p in pagados if (p.get("fecha_pago") or "").startswith(hoy_str))
    pagados_mes = [p for p in pagados if (p.get("fecha_pago") or "").startswith(mes_str)]
    ingresos_mes = sum(float(p.get("monto_total") or 0) for p in pagados_mes)
    ticket_promedio = (sum(float(p.get("monto_total") or 0) for p in pagados) / len(pagados)) if pagados else 0.0

    # Ingresos por día del mes actual
    ingresos_por_dia = [0.0] * dias_mes
    for p in pagados_mes:
        fecha = p.get("fecha_pago") or ""
        try:
            dia = int(fecha[8:10])
            if 1 <= dia <= dias_mes:
                ingresos_por_dia[dia - 1] += float(p.get("monto_total") or 0)
        except (ValueError, IndexError):
            pass
    chart_dias_labels = [str(d) for d in range(1, dias_mes + 1)]
    chart_dias_values = [round(v, 2) for v in ingresos_por_dia]

    # Distribución por método de pago (solo pagados)
    metodo_counts = {m: 0 for m in METODOS_PAGO}
    for p in pagados:
        m = p.get("metodo_pago")
        if m in metodo_counts:
            metodo_counts[m] += 1
    chart_metodo_labels = ["Efectivo", "Tarjeta", "Transferencia"]
    chart_metodo_values = [metodo_counts["efectivo"], metodo_counts["tarjeta"], metodo_counts["transferencia"]]

    # Pagos por estado
    estado_labels_map = {
        "pagado": "Pagado", "pendiente": "Pendiente",
        "verificacion_pendiente": "Verificación pendiente",
        "pendiente_efectivo": "Pendiente efectivo",
        "rechazado": "Rechazado", "cancelado": "Cancelado",
    }
    estado_counts: dict[str, int] = {}
    for p in pagos:
        e = p.get("estado") or "pendiente"
        estado_counts[e] = estado_counts.get(e, 0) + 1
    chart_estado_labels = [estado_labels_map.get(e, e) for e in estado_counts]
    chart_estado_values = list(estado_counts.values())

    pagos_filtrados = _filtrar_pagos(pagos, estado, metodo, fecha_inicio, fecha_fin)

    return templates.TemplateResponse(request, "gerente/reportes_pagos.html", _ctx(request,
        title="Reportes de Pagos", flash=flash, pb_ok=pb_ok,
        ingresos_dia=ingresos_dia, ingresos_mes=ingresos_mes, ticket_promedio=ticket_promedio,
        chart_dias_labels=json.dumps(chart_dias_labels),
        chart_dias_values=json.dumps(chart_dias_values),
        chart_metodo_labels=json.dumps(chart_metodo_labels),
        chart_metodo_values=json.dumps(chart_metodo_values),
        chart_estado_labels=json.dumps(chart_estado_labels),
        chart_estado_values=json.dumps(chart_estado_values),
        pagos=pagos_filtrados,
        estado=estado, metodo=metodo, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
        estados_pago=ESTADOS_PAGO, metodos_pago=METODOS_PAGO,
    ))


@router.get("/reportes/pagos/excel")
def reportes_pagos_excel(
    request: Request,
    estado: str = Query("all"),
    metodo: str = Query("all"),
    fecha_inicio: str = Query(""),
    fecha_fin: str = Query(""),
):
    from datetime import datetime

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    pagos = _filtrar_pagos(_pagos_pb(), estado, metodo, fecha_inicio, fecha_fin)

    BLUE, WHITE, ALT = "1E86BD", "FFFFFF", "D6EDF8"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Pagos"

    ws.merge_cells("A1:H1")
    ws["A1"] = "UrbanBike — Reporte de Transacciones"
    ws["A1"].font = Font(name="Calibri", bold=True, color=BLUE, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    estado_map = {"all": "Todos", "pagado": "Pagado", "pendiente": "Pendiente",
                  "verificacion_pendiente": "Verificación pendiente", "pendiente_efectivo": "Pendiente efectivo",
                  "rechazado": "Rechazado", "cancelado": "Cancelado"}
    metodo_map = {"all": "Todos", "efectivo": "Efectivo", "tarjeta": "Tarjeta", "transferencia": "Transferencia"}
    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Estado: {estado_map.get(estado, estado)}  |  Método: {metodo_map.get(metodo, metodo)}  |  "
        f"Período: {fecha_inicio or '—'} → {fecha_fin or '—'}  |  "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws["A2"].font = Font(name="Calibri", color="64748B", size=9)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    hdr_fill = PatternFill("solid", fgColor=BLUE)
    hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    center   = Alignment(horizontal="center", vertical="center")
    right    = Alignment(horizontal="right",  vertical="center")
    thin_bot = Border(bottom=Side(style="thin", color="E2E8F0"))

    headers = ["Nº comprobante", "Ciclista", "Fecha viaje", "Duración (min)", "Método", "Monto", "Estado", "Confirmado por"]
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
    ws.row_dimensions[4].height = 20

    alt_fill = PatternFill("solid", fgColor=ALT)
    dat_font = Font(name="Calibri", size=10)
    metodo_label = {"efectivo": "Efectivo", "tarjeta": "Tarjeta", "transferencia": "Transferencia"}
    estado_label = {"pagado": "Pagado", "pendiente": "Pendiente", "verificacion_pendiente": "Verificación pendiente",
                    "pendiente_efectivo": "Pendiente efectivo", "rechazado": "Rechazado", "cancelado": "Cancelado"}

    for ri, p in enumerate(pagos, start=5):
        valores = [
            p.get("comprobante_numero") or "—",
            p.get("ciclista_nombre") or "—",
            (p.get("fecha_pago") or "—")[:10],
            int(p.get("duracion_minutos") or 0),
            metodo_label.get(p.get("metodo_pago"), p.get("metodo_pago") or "—"),
            float(p.get("monto_total") or 0),
            estado_label.get(p.get("estado"), p.get("estado") or "—"),
            p.get("confirmado_por_empleado_nombre") or "—",
        ]
        for col, val in enumerate(valores, start=1):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = dat_font
            if col == 4:
                c.number_format = "#,##0"
                c.alignment = right
            elif col == 6:
                c.number_format = '"$"#,##0.00'
                c.alignment = right
            if ri % 2 == 0:
                c.fill = alt_fill
            c.border = thin_bot

    total_row = len(pagos) + 5
    tot_font = Font(name="Calibri", bold=True, size=10, color=BLUE)
    ws.cell(row=total_row, column=1, value=f"Total: {len(pagos)} pagos").font = tot_font
    total_monto = ws.cell(row=total_row, column=6, value=sum(float(p.get("monto_total") or 0) for p in pagos))
    total_monto.font = tot_font
    total_monto.number_format = '"$"#,##0.00'
    total_monto.alignment = right

    widths = [22, 26, 14, 16, 16, 14, 22, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"urbanbike_pagos_{fecha_inicio or 'todos'}_{fecha_fin or 'todos'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


_ROLES_EMPLEADO = ["empleado-operacion", "empleado-mantenimiento", "empleado-vigilancia"]
_ROLES_EMPLEADO_LABELS = {
    "empleado-operacion":     "Operación",
    "empleado-mantenimiento": "Mantenimiento",
    "empleado-vigilancia":    "Vigilancia",
}


def _roles_empleado_map(pb) -> dict[str, str]:
    """slug -> id de rol, solo para los 3 roles de empleado."""
    roles = pb.list_records("roles", per_page=50).get("items", [])
    return {r.get("slug"): r["id"] for r in roles if r.get("slug") in _ROLES_EMPLEADO}


def _empleados_pb() -> list[dict]:
    pb = get_admin_client()
    items = pb.list_records("users", expand="rol", sort="email", per_page=200).get("items", [])
    empleados: list[dict] = []
    for u in items:
        rol_obj = (u.get("expand") or {}).get("rol") or {}
        slug = (rol_obj.get("slug") or "").lower()
        if "empleado" in slug:
            empleados.append({
                "id": u.get("id", ""),
                "nombre": u.get("name") or u.get("email", ""),
                "email": u.get("email", ""),
                "rol": rol_obj.get("nombre") or rol_obj.get("slug") or "Sin rol",
                "rol_slug": slug,
                "verificado": bool(u.get("verified")),
                "activo": bool(u.get("activo")),
                "fecha_registro": (u.get("created") or "")[:10],
            })
    return empleados


@router.get("/empleados", response_class=HTMLResponse)
def empleados(request: Request):
    flash = request.session.pop("flash", None)
    empleados_list: list[dict] = []
    pb_ok = True
    try:
        empleados_list = _empleados_pb()
    except Exception:
        pb_ok = False
    return templates.TemplateResponse(request, "gerente/empleados.html", _ctx(request,
        title="Empleados — Gerente", flash=flash,
        empleados=empleados_list, pb_ok=pb_ok,
        roles_empleado=_ROLES_EMPLEADO, roles_empleado_labels=_ROLES_EMPLEADO_LABELS,
    ))


@router.post("/empleados/crear")
def empleados_crear(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    name: str = Form(""),
    rol_slug: str = Form(...),
):
    if rol_slug not in _ROLES_EMPLEADO:
        return _flash(request, "/gerente/empleados", "error", "Rol no válido. Solo puedes crear empleados.")
    try:
        pb = _pb()
        rol_id = _roles_empleado_map(pb).get(rol_slug)
        if not rol_id:
            return _flash(request, "/gerente/empleados", "error", "No se encontró el rol solicitado.")
        payload: dict = {
            "email": email, "password": password, "passwordConfirm": password,
            "emailVisibility": True, "rol": rol_id, "activo": True,
        }
        if name:
            payload["name"] = name
        pb.create_record("users", payload)
        _log(request, "Crear usuario", f"Empleado creado: {email} ({_ROLES_EMPLEADO_LABELS.get(rol_slug, rol_slug)})")
        try:
            pb.request_verification("users", email)
        except Exception:
            pass
        return _flash(request, "/gerente/empleados", "success", "Empleado creado correctamente.")
    except Exception as e:
        return _flash(request, "/gerente/empleados", "error", str(e))


@router.post("/empleados/{uid}/cambiar-rol")
def empleados_cambiar_rol(request: Request, uid: str, rol_slug: str = Form(...)):
    if rol_slug not in _ROLES_EMPLEADO:
        return _flash(request, "/gerente/empleados", "error", "Rol no válido. Solo puedes asignar roles de empleado.")
    try:
        pb = _pb()
        usuario = pb.get_record("users", uid, expand="rol")
        rol_actual_slug = ((usuario.get("expand") or {}).get("rol") or {}).get("slug", "")
        if rol_actual_slug not in _ROLES_EMPLEADO:
            return _flash(request, "/gerente/empleados", "error",
                          "Solo puedes cambiar el rol de usuarios que ya son empleados.")
        rol_id = _roles_empleado_map(pb).get(rol_slug)
        if not rol_id:
            return _flash(request, "/gerente/empleados", "error", "No se encontró el rol solicitado.")
        pb.update_record("users", uid, {"rol": rol_id})
        _log(request, "Editar usuario",
             f"Rol de {usuario.get('email', uid)} cambiado a {_ROLES_EMPLEADO_LABELS.get(rol_slug, rol_slug)}")
        return _flash(request, "/gerente/empleados", "success", "Rol actualizado.")
    except Exception as e:
        return _flash(request, "/gerente/empleados", "error", str(e))


@router.get("/empleados/excel")
def empleados_excel(request: Request):
    from datetime import datetime

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    empleados_list = _empleados_pb()

    BLUE  = "1E86BD"
    WHITE = "FFFFFF"
    ALT   = "D6EDF8"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados"

    ws.merge_cells("A1:E1")
    ws["A1"] = "UrbanBike — Empleados del Sistema"
    ws["A1"].font = Font(name="Calibri", bold=True, color=BLUE, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total: {len(empleados_list)} empleados"
    ws["A2"].font = Font(name="Calibri", color="64748B", size=9)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    hdr_fill = PatternFill("solid", fgColor=BLUE)
    hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    center   = Alignment(horizontal="center", vertical="center")
    thin_bot = Border(bottom=Side(style="thin", color="E2E8F0"))

    columnas = [
        ("Nombre", "nombre"),
        ("Tipo de empleado", "rol"),
        ("Correo electrónico", "email"),
        ("Estado de cuenta", "verificado"),
        ("Fecha de registro", "fecha_registro"),
    ]

    for col, (label, _) in enumerate(columnas, start=1):
        c = ws.cell(row=4, column=col, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
    ws.row_dimensions[4].height = 20

    alt_fill = PatternFill("solid", fgColor=ALT)
    dat_font = Font(name="Calibri", size=10)

    for ri, emp in enumerate(empleados_list, start=5):
        for col, (_, key) in enumerate(columnas, start=1):
            valor = emp.get(key, "")
            if key == "verificado":
                valor = "Verificado" if valor else "Pendiente"
            c = ws.cell(row=ri, column=col, value=valor)
            c.font = dat_font
            c.border = thin_bot
            if ri % 2 == 0:
                c.fill = alt_fill

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"urbanbike_empleados_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Bicicletas ─────────────────────────────────────────────────────────────────

@router.get("/bicicletas", response_class=HTMLResponse)
def bicicletas_list(request: Request):
    flash = request.session.pop("flash", None)
    items: list = []
    estaciones: list = []
    error: str | None = None
    try:
        items = _pb().list_records("bicicletas", sort="codigo", per_page=500).get("items", [])
    except Exception as e:
        error = str(e)
    try:
        estaciones = _pb().list_records("estaciones", sort="nombre", per_page=200).get("items", [])
    except Exception:
        pass
    return templates.TemplateResponse(request, "gerente/bicicletas.html", _ctx(request,
        title="Bicicletas", items=items, estaciones=estaciones, flash=flash, error=error,
    ))


def _validar_foto(foto: UploadFile | None) -> tuple[bool, str | None]:
    """Valida que la foto sea jpg/png y no supere 3MB. Devuelve (tiene_foto, error)."""
    tiene_foto = foto is not None and foto.filename
    if not tiene_foto:
        return False, None
    if foto.content_type not in ("image/jpeg", "image/png"):
        return True, "La foto debe ser un archivo JPG o PNG."
    if foto.size and foto.size > 3 * 1024 * 1024:
        return True, "La foto no debe superar los 3 MB."
    return True, None


@router.post("/bicicletas/crear")
def bicicletas_crear(
    request: Request,
    codigo: str = Form(...),
    tipo: str = Form("classic_bike"),
    estado: str = Form("disponible"),
    estacion: str = Form(""),
    notas: str = Form(""),
    foto: UploadFile | None = File(None),
):
    tiene_foto, error_foto = _validar_foto(foto)
    if error_foto:
        return _flash(request, "/gerente/bicicletas", "error", error_foto)
    try:
        pb = _pb()
        registro = pb.create_record("bicicletas", {
            "codigo": codigo, "tipo": tipo, "estado": estado,
            "estacion": estacion, "notas": notas,
        })
        if tiene_foto:
            contenido = foto.file.read()
            pb.update_record_with_file("bicicletas", registro["id"], {},
                {"foto": (foto.filename, contenido, foto.content_type)})
        _log(request, "Crear bicicleta", f"Bicicleta registrada: {codigo}")
        return _flash(request, "/gerente/bicicletas", "success", "Bicicleta registrada.")
    except Exception as e:
        return _flash(request, "/gerente/bicicletas", "error", str(e))


@router.post("/bicicletas/{bid}/editar")
def bicicletas_editar(
    request: Request, bid: str,
    codigo: str = Form(""), tipo: str = Form(""),
    estado: str = Form(""), estacion: str = Form(""),
    notas: str = Form(""),
    foto: UploadFile | None = File(None),
):
    tiene_foto, error_foto = _validar_foto(foto)
    if error_foto:
        return _flash(request, "/gerente/bicicletas", "error", error_foto)
    try:
        pb = _pb()
        payload: dict = {"tipo": tipo, "estado": estado, "estacion": estacion, "notas": notas}
        if codigo:
            payload["codigo"] = codigo
        if tiene_foto:
            contenido = foto.file.read()
            pb.update_record_with_file("bicicletas", bid, payload,
                {"foto": (foto.filename, contenido, foto.content_type)})
        else:
            pb.update_record("bicicletas", bid, payload)
        _log(request, "Editar bicicleta", f"Bicicleta actualizada: {codigo or bid}")
        return _flash(request, "/gerente/bicicletas", "success", "Bicicleta actualizada.")
    except Exception as e:
        return _flash(request, "/gerente/bicicletas", "error", str(e))


@router.post("/bicicletas/{bid}/eliminar")
def bicicletas_eliminar(request: Request, bid: str):
    try:
        _pb().delete_record("bicicletas", bid)
        _log(request, "Eliminar bicicleta", f"Bicicleta eliminada (id: {bid})")
        return _flash(request, "/gerente/bicicletas", "success", "Bicicleta eliminada.")
    except Exception as e:
        return _flash(request, "/gerente/bicicletas", "error", str(e))


# ── Estaciones ───────────────────────────────────────────────────────────────

@router.get("/estaciones", response_class=HTMLResponse)
def estaciones_list(request: Request):
    flash = request.session.pop("flash", None)
    items: list = []
    error: str | None = None
    try:
        items = _pb().list_records("estaciones", sort="nombre", per_page=500).get("items", [])
    except Exception as e:
        error = str(e)
    return templates.TemplateResponse(request, "gerente/estaciones.html", _ctx(request,
        title="Estaciones", items=items, flash=flash, error=error,
        estaciones_json=json.dumps(items),
    ))


def _siguiente_codigo_estacion(pb) -> str:
    items = pb.list_records("estaciones", filter='codigo ~ "EST-"', per_page=500).get("items", [])
    maximo = 0
    for e in items:
        partes = (e.get("codigo") or "").split("-")
        if len(partes) == 2 and partes[0] == "EST" and partes[1].isdigit():
            maximo = max(maximo, int(partes[1]))
    return f"EST-{str(maximo + 1).zfill(3)}"


@router.post("/estaciones/crear")
def estaciones_crear(
    request: Request,
    nombre: str = Form(...),
    capacidad: str = Form(""),
    latitud: str = Form(""),
    longitud: str = Form(""),
    activa: str = Form("true"),
):
    try:
        pb = _pb()
        codigo = _siguiente_codigo_estacion(pb)
        payload: dict = {"nombre": nombre, "codigo": codigo, "activa": activa == "true"}
        if capacidad:
            try: payload["capacidad"] = int(capacidad)
            except ValueError: pass
        if latitud:
            try: payload["latitud"] = float(latitud)
            except ValueError: pass
        if longitud:
            try: payload["longitud"] = float(longitud)
            except ValueError: pass
        pb.create_record("estaciones", payload)
        _log(request, "Crear estación", f"Estación creada: {nombre} ({codigo})")
        return _flash(request, "/gerente/estaciones", "success", f"Estación {codigo} creada.")
    except Exception as e:
        return _flash(request, "/gerente/estaciones", "error", str(e))


@router.get("/estaciones/buscar-lugar")
def estaciones_buscar_lugar(request: Request, q: str = Query("")) -> JSONResponse:
    if not q.strip():
        return JSONResponse([])
    params = urllib.parse.urlencode({
        "q": q.strip(),
        "format": "json",
        "limit": 5,
        "countrycodes": "ec",
        "accept-language": "es",
    })
    url = f"https://nominatim.openstreetmap.org/search?{params}"
    req = urllib.request.Request(url, headers={"User-Agent": "UrbanBike-App/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode())
        resultados = [
            {"nombre": item.get("display_name", ""), "lat": float(item["lat"]), "lng": float(item["lon"])}
            for item in data
        ]
        return JSONResponse(resultados)
    except Exception:
        return JSONResponse([])


@router.post("/estaciones/{eid}/editar")
def estaciones_editar(
    request: Request, eid: str,
    nombre: str = Form(""), codigo: str = Form(""),
    capacidad: str = Form(""), latitud: str = Form(""),
    longitud: str = Form(""), activa: str = Form("true"),
):
    try:
        payload: dict = {"activa": activa == "true"}
        if nombre: payload["nombre"] = nombre
        if codigo: payload["codigo"] = codigo
        if capacidad:
            try: payload["capacidad"] = int(capacidad)
            except ValueError: pass
        if latitud:
            try: payload["latitud"] = float(latitud)
            except ValueError: pass
        if longitud:
            try: payload["longitud"] = float(longitud)
            except ValueError: pass
        _pb().update_record("estaciones", eid, payload)
        _log(request, "Editar estación", f"Estación actualizada: {nombre or eid}")
        return _flash(request, "/gerente/estaciones", "success", "Estación actualizada.")
    except Exception as e:
        return _flash(request, "/gerente/estaciones", "error", str(e))


@router.post("/estaciones/{eid}/toggleactiva")
def estaciones_toggleactiva(request: Request, eid: str):
    try:
        pb = _pb()
        est = pb.get_record("estaciones", eid)
        nueva = not bool(est.get("activa"))
        pb.update_record("estaciones", eid, {"activa": nueva})
        _log(request, "Editar estación", f"Estación {est.get('nombre', eid)} marcada como {'activa' if nueva else 'inactiva'}")
        return _flash(request, "/gerente/estaciones", "success", "Estado de la estación actualizado.")
    except Exception as e:
        return _flash(request, "/gerente/estaciones", "error", str(e))


@router.post("/estaciones/{eid}/eliminar")
def estaciones_eliminar(request: Request, eid: str):
    try:
        _pb().delete_record("estaciones", eid)
        _log(request, "Eliminar estación", f"Estación eliminada (id: {eid})")
        return _flash(request, "/gerente/estaciones", "success", "Estación eliminada.")
    except Exception as e:
        return _flash(request, "/gerente/estaciones", "error", str(e))


# ── Tarifas ──────────────────────────────────────────────────────────────────

@router.get("/tarifas", response_class=HTMLResponse)
def tarifas_list(request: Request):
    flash = request.session.pop("flash", None)
    items: list = []
    error: str | None = None
    try:
        items = _pb().list_records("tarifas", sort="tipo_bicicleta", per_page=200).get("items", [])
    except Exception as e:
        error = str(e)
    return templates.TemplateResponse(request, "gerente/tarifas.html", _ctx(request,
        title="Tarifas", items=items, flash=flash, error=error,
    ))


@router.post("/tarifas/crear")
def tarifas_crear(
    request: Request,
    tipo_bicicleta: str = Form(...),
    tipo_usuario: str = Form(...),
    precio_hora: str = Form(...),
    activa: str = Form("true"),
):
    try:
        _pb().create_record("tarifas", {
            "tipo_bicicleta": tipo_bicicleta,
            "tipo_usuario":   tipo_usuario,
            "precio_hora":    float(precio_hora),
            "activa":         activa == "true",
        })
        _log(request, "Crear tarifa", f"Tarifa creada: {tipo_bicicleta} / {tipo_usuario}")
        return _flash(request, "/gerente/tarifas", "success", "Tarifa creada.")
    except Exception as e:
        return _flash(request, "/gerente/tarifas", "error", str(e))


@router.post("/tarifas/{tid}/editar")
def tarifas_editar(
    request: Request, tid: str,
    tipo_bicicleta: str = Form(""), tipo_usuario: str = Form(""),
    precio_hora: str = Form(""), activa: str = Form("true"),
):
    try:
        payload: dict = {"activa": activa == "true"}
        if tipo_bicicleta: payload["tipo_bicicleta"] = tipo_bicicleta
        if tipo_usuario:   payload["tipo_usuario"]   = tipo_usuario
        if precio_hora:
            try: payload["precio_hora"] = float(precio_hora)
            except ValueError: pass
        _pb().update_record("tarifas", tid, payload)
        _log(request, "Editar tarifa", f"Tarifa actualizada (id: {tid})")
        return _flash(request, "/gerente/tarifas", "success", "Tarifa actualizada.")
    except Exception as e:
        return _flash(request, "/gerente/tarifas", "error", str(e))


@router.post("/tarifas/{tid}/toggleactiva")
def tarifas_toggleactiva(request: Request, tid: str):
    try:
        pb = _pb()
        tarifa = pb.get_record("tarifas", tid)
        nueva = not bool(tarifa.get("activa"))
        pb.update_record("tarifas", tid, {"activa": nueva})
        _log(request, "Editar tarifa", f"Tarifa {tarifa.get('tipo_bicicleta', tid)}/{tarifa.get('tipo_usuario', '')} marcada como {'activa' if nueva else 'inactiva'}")
        return _flash(request, "/gerente/tarifas", "success", "Estado de la tarifa actualizado.")
    except Exception as e:
        return _flash(request, "/gerente/tarifas", "error", str(e))


@router.post("/tarifas/{tid}/eliminar")
def tarifas_eliminar(request: Request, tid: str):
    try:
        _pb().delete_record("tarifas", tid)
        _log(request, "Eliminar tarifa", f"Tarifa eliminada (id: {tid})")
        return _flash(request, "/gerente/tarifas", "success", "Tarifa eliminada.")
    except Exception as e:
        return _flash(request, "/gerente/tarifas", "error", str(e))


@router.get("/informe", response_class=HTMLResponse)
def informe(request: Request):
    flash = request.session.pop("flash", None)
    ch_ok = True

    total_viajes = 0
    precio_promedio = 0.0
    ingresos_estimados = 0.0
    top5: list[dict] = []
    tipo_labels: list = []
    tipo_values: list = []
    membresia_labels: list = []
    membresia_values: list = []
    top5_labels: list = []
    top5_values: list = []

    try:
        total_row = ch.query_one("SELECT count() AS total FROM fact_viajes")
        total_viajes = total_row.get("total", 0) if total_row else 0

        top5 = ch.query("""
            SELECT e.nombre_estacion AS nombre, count() AS viajes
            FROM fact_viajes f
            LEFT JOIN dim_estaciones e ON f.id_estacion_inicio = e.id_estacion
            GROUP BY e.nombre_estacion ORDER BY viajes DESC LIMIT 5
        """)
        top5_labels = [str(r.get("nombre") or "N/A") for r in top5]
        top5_values = [r["viajes"] for r in top5]

        tipo_rows = ch.query("""
            SELECT t.nombre AS nombre, count() AS viajes
            FROM fact_viajes f
            LEFT JOIN dim_tipos_bicicleta t ON f.id_tipo_bicicleta = t.id_tipo
            GROUP BY t.nombre
        """)
        tipo_labels = [str(r.get("nombre") or "N/A") for r in tipo_rows]
        tipo_values = [r["viajes"] for r in tipo_rows]

        membresia_rows = ch.query("""
            SELECT m.tipo AS nombre, count() AS viajes
            FROM fact_viajes f
            LEFT JOIN dim_membresia m ON f.id_membresia = m.id_membresia
            GROUP BY m.tipo
        """)
        membresia_labels = [str(r.get("nombre") or "N/A") for r in membresia_rows]
        membresia_values = [r["viajes"] for r in membresia_rows]
    except Exception:
        ch_ok = False

    try:
        pb = get_admin_client()
        tarifas = pb.list_records("tarifas", per_page=200).get("items", [])
        precios = [float(t.get("precio_hora", 0)) for t in tarifas if t.get("precio_hora") is not None]
        if precios:
            precio_promedio = sum(precios) / len(precios)
    except Exception:
        pass

    ingresos_estimados = total_viajes * precio_promedio

    return templates.TemplateResponse(request, "gerente/informe.html", _ctx(request,
        title="Informe General — Gerente", flash=flash, ch_ok=ch_ok,
        total_viajes=total_viajes,
        precio_promedio=precio_promedio,
        ingresos_estimados=ingresos_estimados,
        top5=top5,
        tipo_labels=json.dumps(tipo_labels),
        tipo_values=json.dumps(tipo_values),
        membresia_labels=json.dumps(membresia_labels),
        membresia_values=json.dumps(membresia_values),
        top5_labels=json.dumps(top5_labels),
        top5_values=json.dumps(top5_values),
    ))
