"""Rutas CRUD para la sección administrativa: usuarios, bicicletas, estaciones, tarifas."""

import io
import json
from datetime import datetime, timezone

from fastapi import APIRouter, File, Form, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

from app.db.pocketbase import get_admin_client, PocketBaseError, registrar_auditoria
from app.templating import templates

router = APIRouter(prefix="/admin", tags=["admin"])


def _flash(request: Request, url: str, tipo: str, msg: str) -> RedirectResponse:
    request.session["flash"] = {"type": tipo, "msg": msg}
    return RedirectResponse(url, status_code=302)


def _ctx(request: Request, **extra) -> dict:
    return {"user": getattr(request.state, "user", None), **extra}


def _pb():
    """Admin client; resetea el singleton si el token expiró."""
    import app.db.pocketbase as pbmod
    try:
        return get_admin_client()
    except Exception:
        pbmod._admin_client = None
        return get_admin_client()


_ACCION_TIPO = {"crear": "crear", "editar": "editar", "eliminar": "eliminar"}
_MODULO_PLURAL = {
    "usuario": "usuarios", "bicicleta": "bicicletas",
    "estación": "estaciones", "tarifa": "tarifas",
}


def _log(request: Request, accion: str, detalle: str) -> None:
    """Registra una acción de CRUD en la bitácora de cambios y en la auditoría."""
    user = getattr(request.state, "user", {}) or {}
    try:
        _pb().create_record("bitacora_cambios", {
            "usuario_nombre": user.get("name") or user.get("email", "Admin"),
            "accion":  accion,
            "detalle": detalle,
            "fecha":   datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        })
    except Exception:
        pass

    palabras = accion.lower().split(" ", 1)
    accion_tipo = _ACCION_TIPO.get(palabras[0], "editar")
    modulo = _MODULO_PLURAL.get(palabras[1], "sistema") if len(palabras) > 1 else "sistema"
    registrar_auditoria(
        user.get("pb_token", ""), user.get("id", ""),
        user.get("name") or user.get("email", "Admin"), user.get("email", ""),
        accion_tipo, modulo, detalle, request,
        usuario_rol=user.get("rol_nombre") or user.get("rol_slug", ""),
    )


# ── USUARIOS ──────────────────────────────────────────────────────────────────

@router.get("/usuarios", response_class=HTMLResponse)
def usuarios_list(request: Request):
    flash = request.session.pop("flash", None)
    items: list = []
    roles: list = []
    error: str | None = None
    try:
        pb = _pb()
        items = pb.list_records("users", sort="-created", per_page=100, expand="rol").get("items", [])
    except Exception as e:
        error = str(e)
    try:
        roles = _pb().list_records("roles", per_page=50, sort="nombre").get("items", [])
    except Exception:
        pass
    return templates.TemplateResponse(request, "admin/usuarios.html", _ctx(request,
        title="Usuarios", items=items, roles=roles, flash=flash, error=error,
    ))


@router.post("/usuarios/crear")
def usuarios_crear(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    name: str = Form(""),
    rol: str = Form(""),
):
    try:
        payload: dict = {"email": email, "password": password,
                         "passwordConfirm": password, "emailVisibility": True}
        if name: payload["name"] = name
        if rol:  payload["rol"]  = rol
        pb = _pb()
        pb.create_record("users", payload)
        _log(request, "Crear usuario", f"Usuario creado: {email}")

        try:
            pb.request_verification("users", email)
            return _flash(request, "/admin/usuarios", "success",
                          "Usuario creado correctamente. Se envió un correo de verificación.")
        except Exception:
            return _flash(request, "/admin/usuarios", "success",
                          "Usuario creado correctamente, pero no se pudo enviar el correo de verificación.")
    except Exception as e:
        return _flash(request, "/admin/usuarios", "error", str(e))


@router.post("/usuarios/{uid}/editar")
def usuarios_editar(
    request: Request, uid: str,
    name: str = Form(""), rol: str = Form(""),
):
    try:
        payload: dict = {"rol": rol}
        if name: payload["name"] = name
        _pb().update_record("users", uid, payload)
        _log(request, "Editar usuario", f"Usuario actualizado (id: {uid})")
        return _flash(request, "/admin/usuarios", "success", "Usuario actualizado.")
    except Exception as e:
        return _flash(request, "/admin/usuarios", "error", str(e))


@router.post("/usuarios/{uid}/eliminar")
def usuarios_eliminar(request: Request, uid: str):
    try:
        _pb().delete_record("users", uid)
        _log(request, "Eliminar usuario", f"Usuario eliminado (id: {uid})")
        return _flash(request, "/admin/usuarios", "success", "Usuario eliminado.")
    except Exception as e:
        return _flash(request, "/admin/usuarios", "error", str(e))


# ── BICICLETAS ────────────────────────────────────────────────────────────────

@router.get("/bicicletas", response_class=HTMLResponse)
def bicicletas_list(request: Request):
    flash = request.session.pop("flash", None)
    items: list = []
    estaciones: list = []
    error: str | None = None
    try:
        items = _pb().list_records("bicicletas", sort="codigo", per_page=100).get("items", [])
    except Exception as e:
        error = str(e)
    try:
        estaciones = _pb().list_records("estaciones", sort="nombre", per_page=200).get("items", [])
    except Exception:
        pass
    return templates.TemplateResponse(request, "admin/bicicletas.html", _ctx(request,
        title="Bicicletas", items=items, estaciones=estaciones, flash=flash, error=error,
    ))


def _validar_foto(foto: UploadFile | None) -> tuple[bool, str | None]:
    """Valida que la foto sea jpg/png y no supere 3MB. Devuelve (tiene_foto, error)."""
    tiene_foto = foto is not None and foto.filename
    if not tiene_foto:
        return False, None
    if foto.content_type not in ("image/jpeg", "image/png"):
        return True, "La foto debe ser un archivo JPG o PNG."
    if foto.size and foto.size > 3 * 1024 * 1024:
        return True, "La foto no debe superar los 3 MB."
    return True, None


@router.post("/bicicletas/crear")
def bicicletas_crear(
    request: Request,
    codigo: str = Form(...),
    tipo: str = Form("classic_bike"),
    estado: str = Form("disponible"),
    estacion: str = Form(""),
    notas: str = Form(""),
    foto: UploadFile | None = File(None),
):
    tiene_foto, error_foto = _validar_foto(foto)
    if error_foto:
        return _flash(request, "/admin/bicicletas", "error", error_foto)
    try:
        payload = {"codigo": codigo, "tipo": tipo, "estado": estado,
                   "estacion": estacion, "notas": notas}
        registro = _pb().create_record("bicicletas", payload)
        if tiene_foto:
            contenido = foto.file.read()
            _pb().update_record_with_file("bicicletas", registro["id"], {},
                {"foto": (foto.filename, contenido, foto.content_type)})
        _log(request, "Crear bicicleta", f"Bicicleta registrada: {codigo}")
        return _flash(request, "/admin/bicicletas", "success", "Bicicleta registrada.")
    except Exception as e:
        return _flash(request, "/admin/bicicletas", "error", str(e))


@router.post("/bicicletas/{bid}/editar")
def bicicletas_editar(
    request: Request, bid: str,
    codigo: str = Form(""), tipo: str = Form(""),
    estado: str = Form(""), estacion: str = Form(""),
    notas: str = Form(""),
    foto: UploadFile | None = File(None),
):
    tiene_foto, error_foto = _validar_foto(foto)
    if error_foto:
        return _flash(request, "/admin/bicicletas", "error", error_foto)
    try:
        payload: dict = {"tipo": tipo, "estado": estado,
                         "estacion": estacion, "notas": notas}
        if codigo: payload["codigo"] = codigo
        if tiene_foto:
            contenido = foto.file.read()
            _pb().update_record_with_file("bicicletas", bid, payload,
                {"foto": (foto.filename, contenido, foto.content_type)})
        else:
            _pb().update_record("bicicletas", bid, payload)
        _log(request, "Editar bicicleta", f"Bicicleta actualizada: {codigo or bid}")
        return _flash(request, "/admin/bicicletas", "success", "Bicicleta actualizada.")
    except Exception as e:
        return _flash(request, "/admin/bicicletas", "error", str(e))


@router.post("/bicicletas/{bid}/eliminar")
def bicicletas_eliminar(request: Request, bid: str):
    try:
        _pb().delete_record("bicicletas", bid)
        _log(request, "Eliminar bicicleta", f"Bicicleta eliminada (id: {bid})")
        return _flash(request, "/admin/bicicletas", "success", "Bicicleta eliminada.")
    except Exception as e:
        return _flash(request, "/admin/bicicletas", "error", str(e))


# ── ESTACIONES ────────────────────────────────────────────────────────────────

@router.get("/estaciones", response_class=HTMLResponse)
def estaciones_list(request: Request):
    flash = request.session.pop("flash", None)
    items: list = []
    error: str | None = None
    try:
        items = _pb().list_records("estaciones", sort="nombre", per_page=100).get("items", [])
    except Exception as e:
        error = str(e)
    return templates.TemplateResponse(request, "admin/estaciones.html", _ctx(request,
        title="Estaciones", items=items, flash=flash, error=error,
        estaciones_json=json.dumps(items),
    ))


@router.post("/estaciones/crear")
def estaciones_crear(
    request: Request,
    nombre: str = Form(...),
    codigo: str = Form(""),
    capacidad: str = Form(""),
    latitud: str = Form(""),
    longitud: str = Form(""),
    activa: str = Form("true"),
):
    try:
        payload: dict = {"nombre": nombre, "activa": activa == "true"}
        if codigo: payload["codigo"] = codigo
        if capacidad:
            try: payload["capacidad"] = int(capacidad)
            except ValueError: pass
        if latitud:
            try: payload["latitud"] = float(latitud)
            except ValueError: pass
        if longitud:
            try: payload["longitud"] = float(longitud)
            except ValueError: pass
        _pb().create_record("estaciones", payload)
        _log(request, "Crear estación", f"Estación creada: {nombre}")
        return _flash(request, "/admin/estaciones", "success", "Estación creada.")
    except Exception as e:
        return _flash(request, "/admin/estaciones", "error", str(e))


@router.post("/estaciones/{eid}/editar")
def estaciones_editar(
    request: Request, eid: str,
    nombre: str = Form(""), codigo: str = Form(""),
    capacidad: str = Form(""), latitud: str = Form(""),
    longitud: str = Form(""), activa: str = Form("true"),
):
    try:
        payload: dict = {"activa": activa == "true"}
        if nombre: payload["nombre"] = nombre
        if codigo: payload["codigo"] = codigo
        if capacidad:
            try: payload["capacidad"] = int(capacidad)
            except ValueError: pass
        if latitud:
            try: payload["latitud"] = float(latitud)
            except ValueError: pass
        if longitud:
            try: payload["longitud"] = float(longitud)
            except ValueError: pass
        _pb().update_record("estaciones", eid, payload)
        _log(request, "Editar estación", f"Estación actualizada: {nombre or eid}")
        return _flash(request, "/admin/estaciones", "success", "Estación actualizada.")
    except Exception as e:
        return _flash(request, "/admin/estaciones", "error", str(e))


@router.post("/estaciones/{eid}/eliminar")
def estaciones_eliminar(request: Request, eid: str):
    try:
        _pb().delete_record("estaciones", eid)
        _log(request, "Eliminar estación", f"Estación eliminada (id: {eid})")
        return _flash(request, "/admin/estaciones", "success", "Estación eliminada.")
    except Exception as e:
        return _flash(request, "/admin/estaciones", "error", str(e))


# ── TARIFAS ───────────────────────────────────────────────────────────────────

@router.get("/tarifas", response_class=HTMLResponse)
def tarifas_list(request: Request):
    flash = request.session.pop("flash", None)
    items: list = []
    error: str | None = None
    try:
        items = _pb().list_records("tarifas", sort="tipo_bicicleta", per_page=100).get("items", [])
    except Exception as e:
        error = str(e)
    return templates.TemplateResponse(request, "admin/tarifas.html", _ctx(request,
        title="Tarifas", items=items, flash=flash, error=error,
    ))


@router.post("/tarifas/crear")
def tarifas_crear(
    request: Request,
    tipo_bicicleta: str = Form(...),
    tipo_usuario: str = Form(...),
    precio_hora: str = Form(...),
    activa: str = Form("true"),
):
    try:
        _pb().create_record("tarifas", {
            "tipo_bicicleta": tipo_bicicleta,
            "tipo_usuario":   tipo_usuario,
            "precio_hora":    float(precio_hora),
            "activa":         activa == "true",
        })
        _log(request, "Crear tarifa", f"Tarifa creada: {tipo_bicicleta} / {tipo_usuario}")
        return _flash(request, "/admin/tarifas", "success", "Tarifa creada.")
    except Exception as e:
        return _flash(request, "/admin/tarifas", "error", str(e))


@router.post("/tarifas/{tid}/editar")
def tarifas_editar(
    request: Request, tid: str,
    tipo_bicicleta: str = Form(""), tipo_usuario: str = Form(""),
    precio_hora: str = Form(""), activa: str = Form("true"),
):
    try:
        payload: dict = {"activa": activa == "true"}
        if tipo_bicicleta: payload["tipo_bicicleta"] = tipo_bicicleta
        if tipo_usuario:   payload["tipo_usuario"]   = tipo_usuario
        if precio_hora:
            try: payload["precio_hora"] = float(precio_hora)
            except ValueError: pass
        _pb().update_record("tarifas", tid, payload)
        _log(request, "Editar tarifa", f"Tarifa actualizada (id: {tid})")
        return _flash(request, "/admin/tarifas", "success", "Tarifa actualizada.")
    except Exception as e:
        return _flash(request, "/admin/tarifas", "error", str(e))


@router.post("/tarifas/{tid}/eliminar")
def tarifas_eliminar(request: Request, tid: str):
    try:
        _pb().delete_record("tarifas", tid)
        _log(request, "Eliminar tarifa", f"Tarifa eliminada (id: {tid})")
        return _flash(request, "/admin/tarifas", "success", "Tarifa eliminada.")
    except Exception as e:
        return _flash(request, "/admin/tarifas", "error", str(e))


# ── BITÁCORA DE CAMBIOS ───────────────────────────────────────────────────────

@router.get("/bitacora", response_class=HTMLResponse)
def bitacora_list(request: Request):
    flash = request.session.pop("flash", None)
    items: list = []
    error: str | None = None
    try:
        items = _pb().list_records("bitacora_cambios", sort="-fecha", per_page=50).get("items", [])
    except Exception as e:
        error = str(e)
    return templates.TemplateResponse(request, "admin/bitacora.html", _ctx(request,
        title="Bitácora de Cambios", items=items, flash=flash, error=error,
    ))


# ── AUDITORÍA ─────────────────────────────────────────────────────────────────

def _auditoria_filtro(accion: str, modulo: str, usuario: str) -> str:
    partes: list[str] = []
    if accion:
        partes.append(f'accion = "{accion}"')
    if modulo:
        partes.append(f'modulo = "{modulo}"')
    if usuario:
        u = usuario.replace('"', '')
        partes.append(f'usuario_nombre ~ "{u}"')
    return " && ".join(partes)


@router.get("/auditoria", response_class=HTMLResponse)
def auditoria_list(
    request: Request,
    accion:  str = Query(""),
    modulo:  str = Query(""),
    usuario: str = Query(""),
):
    flash = request.session.pop("flash", None)
    items: list = []
    error: str | None = None
    try:
        items = _pb().list_records(
            "auditoria",
            filter=_auditoria_filtro(accion, modulo, usuario),
            sort="-fecha", per_page=100,
        ).get("items", [])
    except Exception as e:
        error = str(e)
    return templates.TemplateResponse(request, "admin/auditoria.html", _ctx(request,
        title="Auditoría", items=items, flash=flash, error=error,
        f_accion=accion, f_modulo=modulo, f_usuario=usuario,
    ))


@router.get("/auditoria/excel")
def auditoria_excel(
    request: Request,
    accion:  str = Query(""),
    modulo:  str = Query(""),
    usuario: str = Query(""),
):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    items = _pb().list_records(
        "auditoria",
        filter=_auditoria_filtro(accion, modulo, usuario),
        sort="-fecha", per_page=100,
    ).get("items", [])

    BLUE  = "1E86BD"
    WHITE = "FFFFFF"
    ALT   = "D6EDF8"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoría"

    cols = ["Fecha/Hora", "Usuario", "Email", "Rol", "Acción", "Módulo", "Detalle", "IP"]

    ws.merge_cells(f"A1:{chr(64 + len(cols))}1")
    ws["A1"] = "UrbanBike — Registro de Auditoría"
    ws["A1"].font = Font(name="Calibri", bold=True, color=BLUE, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{chr(64 + len(cols))}2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total: {len(items)} registros"
    ws["A2"].font = Font(name="Calibri", color="64748B", size=9)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    hdr_fill = PatternFill("solid", fgColor=BLUE)
    hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    center   = Alignment(horizontal="center", vertical="center")
    thin_bot = Border(bottom=Side(style="thin", color="E2E8F0"))

    for col, label in enumerate(cols, start=1):
        c = ws.cell(row=4, column=col, value=label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
    ws.row_dimensions[4].height = 20

    alt_fill = PatternFill("solid", fgColor=ALT)
    dat_font = Font(name="Calibri", size=10)
    keys = ["fecha", "usuario_nombre", "usuario_email", "usuario_rol", "accion", "modulo", "detalle", "ip_cliente"]

    for ri, item in enumerate(items, start=5):
        for col, key in enumerate(keys, start=1):
            valor = item.get(key, "")
            if key == "fecha":
                valor = str(valor).replace("T", " ").replace("Z", "")
            c = ws.cell(row=ri, column=col, value=valor)
            c.font = dat_font
            c.border = thin_bot
            if ri % 2 == 0:
                c.fill = alt_fill

    anchos = [20, 26, 32, 18, 12, 16, 50, 16]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i)].width = ancho
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"urbanbike_auditoria_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── REPORTES ──────────────────────────────────────────────────────────────────

@router.get("/reportes", response_class=HTMLResponse)
def reportes(request: Request):
    flash = request.session.pop("flash", None)
    error: str | None = None

    rol_labels: list = []
    rol_values: list = []
    bici_labels: list = []
    bici_values: list = []
    est_labels: list = []
    est_values: list = []
    total_usuarios = total_bicis = total_estaciones = 0

    try:
        pb = _pb()

        usuarios = pb.list_records("users", expand="rol", per_page=500).get("items", [])
        rol_counts: dict[str, int] = {}
        for u in usuarios:
            rol_obj = (u.get("expand") or {}).get("rol") or {}
            nombre = rol_obj.get("nombre") or rol_obj.get("slug") or "Sin rol"
            rol_counts[nombre] = rol_counts.get(nombre, 0) + 1
        rol_labels = list(rol_counts.keys())
        rol_values = list(rol_counts.values())
        total_usuarios = len(usuarios)

        bicis = pb.list_records("bicicletas", per_page=500).get("items", [])
        bici_counts = {"disponible": 0, "en_uso": 0, "mantenimiento": 0, "retirada": 0}
        for b in bicis:
            estado = b.get("estado", "disponible")
            bici_counts[estado] = bici_counts.get(estado, 0) + 1
        bici_labels = ["Disponible", "En uso", "Mantenimiento", "Retirada"]
        bici_values = [bici_counts["disponible"], bici_counts["en_uso"],
                       bici_counts["mantenimiento"], bici_counts["retirada"]]
        total_bicis = len(bicis)

        # El dataset es exclusivamente de Nueva York (no existe campo "ciudad" en
        # estaciones); se reporta su distribución por estado operativo.
        estaciones = pb.list_records("estaciones", per_page=500).get("items", [])
        activas = sum(1 for e in estaciones if e.get("activa"))
        inactivas = len(estaciones) - activas
        est_labels = ["Activas", "Inactivas"]
        est_values = [activas, inactivas]
        total_estaciones = len(estaciones)

    except Exception as e:
        error = str(e)

    return templates.TemplateResponse(request, "admin/reportes.html", _ctx(request,
        title="Reportes", flash=flash, error=error,
        total_usuarios=total_usuarios, total_bicis=total_bicis, total_estaciones=total_estaciones,
        rol_labels=json.dumps(rol_labels), rol_values=json.dumps(rol_values),
        bici_labels=json.dumps(bici_labels), bici_values=json.dumps(bici_values),
        est_labels=json.dumps(est_labels), est_values=json.dumps(est_values),
    ))
