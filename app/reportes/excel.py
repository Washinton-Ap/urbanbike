"""Helper único de exportación a Excel — reemplaza las 4 copias duplicadas
que existían en gerente.py (reportes, reportes/pagos, empleados) y admin.py
(auditoría). Mismo esquema de color que ya usaban: encabezado azul, filas
alternadas.
"""

import io
from datetime import datetime

import openpyxl
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.reportes.comun import ALT, BLUE, GRID, MUTED, WHITE, ColumnaReporte, ReporteVacioError

FORMATOS_NUMERO = {
    "texto": None,
    "entero": "#,##0",
    "decimal1": "0.0",
    "decimal2": "0.00",
    "moneda": '"$"#,##0.00',
}


def generar_excel_reporte(
    *,
    titulo: str,
    subtitulo: str,
    columnas: list[ColumnaReporte],
    filas: list[list],
    nombre_archivo: str,
    nombre_hoja: str = "Reporte",
    fila_total: list | None = None,
) -> StreamingResponse:
    """Genera un .xlsx con membrete UrbanBike, encabezado azul, filas
    alternadas y fila de totales opcional. `filas` y `fila_total` deben
    traer los valores ya en el orden de `columnas` (una celda `None` en
    `fila_total` se deja vacía).
    """
    if not filas:
        raise ReporteVacioError("No hay datos para exportar.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nombre_hoja

    n_cols = len(columnas)
    ultima_col = get_column_letter(n_cols)

    ws.merge_cells(f"A1:{ultima_col}1")
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Calibri", bold=True, color=BLUE, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{ultima_col}2")
    ws["A2"] = f"{subtitulo}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", color=MUTED, size=9)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    hdr_fill = PatternFill("solid", fgColor=BLUE)
    hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin_bot = Border(bottom=Side(style="thin", color=GRID))

    FILA_ENCABEZADO = 4
    for col, columna in enumerate(columnas, start=1):
        c = ws.cell(row=FILA_ENCABEZADO, column=col, value=columna.titulo)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
    ws.row_dimensions[FILA_ENCABEZADO].height = 20

    alt_fill = PatternFill("solid", fgColor=ALT)
    dat_font = Font(name="Calibri", size=10)

    fila_inicio_datos = FILA_ENCABEZADO + 1
    for ri, fila in enumerate(filas, start=fila_inicio_datos):
        for col, (columna, valor) in enumerate(zip(columnas, fila), start=1):
            c = ws.cell(row=ri, column=col, value=valor)
            c.font = dat_font
            c.border = thin_bot
            fmt = FORMATOS_NUMERO.get(columna.formato)
            if fmt:
                c.number_format = fmt
            c.alignment = Alignment(
                horizontal=columna.alineacion_efectiva(), vertical="center",
                wrap_text=(columna.formato == "texto"),
            )
            if ri % 2 == 0:
                c.fill = alt_fill

    if fila_total is not None:
        total_row = fila_inicio_datos + len(filas)
        tot_font = Font(name="Calibri", bold=True, size=10, color=BLUE)
        for col, (columna, valor) in enumerate(zip(columnas, fila_total), start=1):
            if valor is None:
                continue
            c = ws.cell(row=total_row, column=col, value=valor)
            c.font = tot_font
            fmt = FORMATOS_NUMERO.get(columna.formato)
            if fmt:
                c.number_format = fmt
            c.alignment = Alignment(
                horizontal=columna.alineacion_efectiva(),
                wrap_text=(columna.formato == "texto"),
            )

    for i, columna in enumerate(columnas, start=1):
        ws.column_dimensions[get_column_letter(i)].width = columna.ancho
    ws.freeze_panes = f"A{fila_inicio_datos}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )
